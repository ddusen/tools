#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
import xlrd
import xlwt
import numpy
import math

from openpyxl import load_workbook
from mysql import query, query_one, save

black_enterprise_list = [u'湖北新亿通机车配件有限公司',
                        u'湖北晶洋科技股份有限公司',
                        u'襄阳市湖北银城纺织股份有限公司',
                        u'襄阳市宜城安达特种水泥有限公司',
                        u'襄阳市襄阳美利信科技有限责任公司',
                        u'襄阳中车电机技术有限公司',
                        u'湖北台基半导体股份有限公司',
                        u'枣阳市龙居建材制品有限公司',
                        u'枣阳福星摩擦材料有限公司',
                        u'湖北泰和电气有限公司',
                        u'湖北金环新材料科技有限公司',
                        u'武汉嘉仪通科技有限公司',
                        u'武汉普惠海洋光电技术有限公司',
                        u'武汉中仪物联技术股份有限公司',
                        u'3380武汉首家厨房',
                        u'3490武汉重工（浆轴成套',
                        u'武汉重工铸锻有限责任公司',
                        u'长飞光纤',
                        u'湖北溢德数字文化股份有限公司',
                        u'湖北华特红旗电缆有限公司',
                        u'武汉汉麻生物科技有限公司',
                        u'天津科瑞嘉（湖北）机电技术有限公司',
                        u'崇阳县昌华实业有限公司',
                        u'宇恒电气（湖北）有限公司',
                        u'澳森木业赤壁有限责任公司',
                        u'湖北通山赛钻石英建材有限公司',
                        u'湖北红旗中益特种线缆有限责任公司',
                        u'乐星红旗电缆（湖北）有限公司',
                        u'襄阳市襄阳五二五泵业有限公司',
                        u'襄阳市湖北环宇车灯有限公司',
                        u'襄阳市湖北铭航新能源科技有限公司',
                        u'襄阳双金电气成套设备有限公司',
                        u'襄阳市宜城市天舒纺织有限公司',
                        u'际华三五零九纺织有限公司',
                        u'湖北紫竹林科技发展有限公司',
                        u'湖北福星新材料科技有限公司钢丝绳',
                        u'3140武汉重工（钢管）',
                        u'湖北白兆山水泥有限公司',
                        u'湖北爱仕达电器有限公司',
                        u'湖北三江船艇科技有限公司',
                        u'黄石市人本轴承有限公司',
                        u'黄石市富宏智能科技股份有限公司',
                        u'湖北首开机械有限公司',
                        u'湖北远大生命科学与技术有限责任公司',
                        u'黄石市华新水泥股份有限公司',
                        u'丹江口开泰激素',
                        u'汉江丹江口铝业有限责任公司',
                        u'十堰飞纳科科技有限公司',
                        u'十堰市明诚线缆有限公司',
                        u'湖北鑫威机械设备股份有限公司',
                        u'湖北永晟塑料管业有限责任公司',
                        u'湖北扬子江泵业有限责任公司',
                        u'湖北中油科昊机械制造有限公司',
                        u'新动力电机（荆州）有限公司',
                        u'荆州市天宇汽车配件有限公司',
                        u'湖北晶星科技股份有限公司',
                        u'湖北广净环保催化剂有限公司',
                        u'湖北农谷环保科技有限公司',
                        u'襄阳市湖北骆驼物流有限公司',
                        u'武汉钢铁集团鄂城钢铁有限责任公司',
                        u'湖北西迈电气设备有限公司',
                        u'鄂州市金锋超硬材料有限公司',
                        u'湖北弘山塑业有限公司',
                        u'湖北恒颖超科技有限公司',
                        u'黄石市大冶有色金属有限责任公司',
                        u'黄石市东贝电器股份有限公司',
                        u'湖北三丰智能输送装备股份有限公司',
                        u'湖北孚龙科技股份有限公司',
                        u'黄石市新兴管业有限公司',
                        u'湖北大帆金属制品有限公司',
                        u'亚东水泥黄冈亚东水泥有限公司',
                        u'湖北华龙车灯有限公司',
                        u'湖北凯龙化工集团股份有限公司',
                        u'湖北鄂电萃宇电缆有限公司',
                        u'湖北金汉江精制棉有限公司',
                        u'湖北钟格塑料管有限公司',
                        u'湖北鄂电德力电气有限公司',
                        u'沙洋武汉富泰革基布有限公司',
                        u'监利县杨林山水泥有限公司',
                        u'荆州亮诚科技股份有限公司',
                        u'湖北美标汽车制冷系统有限公司',
                        u'荆州劲力建材有限公司',
                        u'黄冈市黄商集团股份有限公司',
                        u'湖北富迪实业股份有限公司']

def get_file_path():
    all_path = os.walk('/mnt/hgfs/Data/work/survey/汇总/')

    xls_files = []
    xlsx_files = []
    for path in all_path:
        for file in path[-1]:
            if file.find(".xlsx") != -1:
                file = "%s/%s" % (path[0], file,)
                xlsx_files.append(file)
            elif file.find(".xls") != -1:
                file = "%s/%s" % (path[0], file,)
                xls_files.append(file)

    return (xls_files, xlsx_files,)

def is_text_contain(sheet_name):
    enterprise_name = sheet_name_to_enterprise_name(sheet_name)
    flag = False
    for name in black_enterprise_list:
        flag = name.find(enterprise_name) != -1 or enterprise_name.find(name) != -1
    return flag

def extract_category(file_path):
    return unicode("%s-%s" % (file_path.split('/')[-2], file_path.split('/')[-1].split('.xls')[0]), "utf-8")

def sheet_name_to_enterprise_name(sheet_name):
    enterprise_name = sheet_name.split(u"、")[1] if sheet_name.find(u"、") != -1 else sheet_name
    return enterprise_name

def enterprise_news_score(sheet_name):
    enterprise_name = sheet_name_to_enterprise_name(sheet_name)
    positive_count = query_one(sql=u'SELECT COUNT(*) FROM `base_news` WHERE `status` = 1 AND `keyword` LIKE "%%%s%%"', list1=(enterprise_name, ))
    negative_count = query_one(sql=u'SELECT COUNT(*) FROM `base_news` WHERE `status` = -1 AND `keyword` LIKE "%%%s%%"', list1=(enterprise_name, ))
    score = 7
    score += positive_count.get('COUNT(*)')
    score -= negative_count.get('COUNT(*)')

    if score < 0 :
        score = 0
    elif score > 10:
        score = 10

    return score

def average(score_list, sheet_name):
    avg = numpy.mean(score_list, axis=0)
    avg = round(numpy.mean([enterprise_news_score(sheet_name), avg], axis=0), 2) if not math.isnan(avg) else u'无评分'
    return avg

def handle_xls(xls_path):
    workbook = xlrd.open_workbook(xls_path)

    category = extract_category(xls_path)

    # get sheel
    sheet_names = filter(lambda x: x.find(u'heet') == -1, workbook.sheet_names())

    # accoding to sheel get sheel content
    data = {}
    for sheet_name in sheet_names:
        result =u''
        if is_text_contain(sheet_name):
            # result = u'此企业为问题企业，无调查评分。互联网分数是：%s' % enterprise_news_score(sheet_name)
            continue
        else:
            sheet = workbook.sheet_by_name(sheet_name)

            rownum = sheet.nrows  # row number
            colnum = sheet.ncols  # col number
            colnames = sheet.row_values(0)
            score_list = []
            for i in xrange(1, rownum):
                score_list += sheet.row_values(i)

            score_list = filter(lambda x: x != u'' and isinstance(x, float), score_list)
            result = average(score_list, sheet_name)

        data['%s : %s' % (category, sheet_name)] = result

    return data


def handle_xlsx(xlsx_path):
    wb = load_workbook(filename=xlsx_path)

    category = extract_category(xlsx_path)

    sheet_names = filter(lambda x: x.find(u'heet') == -1,wb.get_sheet_names())  # 获取所有表格(worksheet)的名字

    data = {}
    for sheet_name in sheet_names:
        result =u''
        if is_text_contain(sheet_name):
            # result = u'此企业为问题企业，无调查评分。互联网分数是：%s' % enterprise_news_score(sheet_name)
            continue
        else:
            ws = wb.get_sheet_by_name(sheet_name)  # 获取特定的 worksheet

            # 获取表格所有行和列，两者都是可迭代的
            rows = ws.rows
            columns = ws.columns

            # 行迭代
            score_list = []
            for index, row in enumerate(rows):
                if index > 0:
                    for col in row:
                        x = col.value
                        if isinstance(x, long):
                            score_list.append(x)
            result = average(score_list, sheet_name)
        data['%s : %s' % (category, sheet_name)] = result

    return data


def write_xls(data):
    file = xlwt.Workbook()                # 注意这里的Workbook首字母是大写
    table = file.add_sheet('statistics', cell_overwrite_ok=True)

    flag = 0
    for d in data:
        for k, v in d.items():
            flag += 1
            table.write(flag, 0, k)
            table.write(flag, 1, v)

    # 保存文件
    file.save('/home/sdu/Project/tools/code/survey/survey_statistics.xls')


def main():
    data = []
    xls_list = get_file_path()[0]
    for xls_path in xls_list:
        data.append(handle_xls(xls_path))

    xlsx_list = get_file_path()[1]
    for xlsx_path in xlsx_list:
        data.append(handle_xlsx(xlsx_path))

    write_xls(data)

if __name__ == '__main__':
    main()
