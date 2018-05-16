import xlsxwriter
import openpyxl

from xlutils.copy import copy
from xlrd import open_workbook, xldate_as_tuple
from xlwt import Workbook
from io import BytesIO
from datetime import datetime


def write_xlsx(filepath, data):
    #data like this
    # data = (
    #     ['Rent', 1000],
    #     ['Gas',   100],
    #     ['Food',  300],
    #     ['Gym',    50],
    # )

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0    
    # Iterate over the data and write it out row by row.
    for r_index, r_data in enumerate(data):
        for c_index, c_data in enumerate(r_data):
            worksheet.write(r_index, c_index, c_data)

    workbook.close()


def append_xls(data):
    #data like this
    # data = (
    #     ['Rent', 1000],
    #     ['Gas',   100],
    #     ['Food',  300],
    #     ['Gym',    50],
    # )

    # Create a workbook and add a worksheet.
    try:
        rexcel = open_workbook("inspections.xls") # 用wlrd提供的方法读取一个excel文件
    except Exception as e:
        print("%s\n%s" % (e, "初始化新文件...", ))        
        workbook = Workbook()
        worksheet = workbook.add_sheet('sheet1')
        workbook.save('inspections.xls')
        rexcel = open_workbook("inspections.xls")

    rows = rexcel.sheets()[0].nrows # 用wlrd提供的方法获得现在已有的行数
    excel = copy(rexcel)# 用xlutils提供的copy方法将xlrd的对象转化为xlwt的对象
    table = excel.get_sheet(0) # 用xlwt对象的方法获得要操作的sheet

    row = rows
    for r_index, r_data in enumerate(data):
        for c_index, c_data in enumerate(r_data):
            table.write(row + r_index, c_index, c_data)# xlwt对象的写方法，参数分别是行、列、值

    excel.save("inspections.xls") # xlwt对象的保存方法，这时便覆盖掉了原来的excel


def read_xlsx(filename):
    wb = openpyxl.load_workbook(filename=filename, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.rows:
        for cell in row:
            pass
    

def read_by_openpyxl(filename=None, file_contents=None, title={}):
    '''
    title like this:
    {
        'GUID': 0, 
        '标题': 0, 
        'URL': 0, 
        '发布时间': 0, 
        '发布媒体': 0, 
        '风险程度': 0, 
        '地域': 0, 
        '类别': 0, 
    }
    '''
    # xlsx_book = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
    xlsx_book = openpyxl.load_workbook(filename=filename, read_only=True)
    sheet = xlsx_book.active
    rows = sheet.rows

    data = []
    for index, row in enumerate(rows):
        if index == 0:
            line = [cell.value for cell in row]
            for k in title.keys():
                title[k] = line.index(k)
        else:
            data2 = {}
            for k, v in title.items():
                data2[k] = sheet.cell(row=index, column=v).value
            data.append(data2)

    return data


def read_by_xlrd(filename=None, file_contents=None, title={}):
    '''
    title like this:
    {
        'GUID': 0, 
        '标题': 0, 
        'URL': 0, 
        '发布时间': 0, 
        '发布媒体': 0, 
        '风险程度': 0, 
        '地域': 0, 
        '类别': 0, 
    }
    '''
    wb = open_workbook(filename=filename, file_contents=file_contents)
    sheet = wb.sheet_by_index(0)
    rows = sheet.nrows
    cols = sheet.ncols

    rvalues = sheet.row_values(0)
    for k in title.keys():
        title[k] = rvalues.index(k)

    # DATA FORMAT
    def format(i, j):
        ctype = sheet.cell(i, j).ctype  # 表格的数据类型
        cell = sheet.cell_value(i, j)
        if ctype == 2 and cell % 1 == 0:  # 如果是整形
            cell = int(cell)
        elif ctype == 3: # 转成datetime对象
            cell = datetime(*xldate_as_tuple(cell, 0))
        elif ctype == 4:
            cell = True if cell == 1 else False

        return cell

    data = []
    for i in range(1, rows):
        data2 = {}
        for k, v in title.items():
            data2[k] = format(i, v)
        data.append(data2)

    return data


if __name__ == '__main__':
    pass
    # filename = '/mnt/f/FileRecv/model.xlsx'
    # filename = '/mnt/f/FileRecv/model2.xlsx'
    # read_xlsx(filename)
    # data = (
    #     ['Rent', 1000],
    #     ['Gas',   100],
    #     ['Food',  300],
    #     ['Gym',    50],
    # )
    # write_xlsx('/home/sdu/Documents/tools/python3/utils/office/test.xlsx', data)
    # append_xls(data)