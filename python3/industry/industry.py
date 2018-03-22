import os, sys
sys.path.append(os.getcwd())

from openpyxl import load_workbook

from utils.db.mysql import query, query_one, save


def db_config():
    return {
        'host':'gz-cdb-ko3zdkzs.sql.tencentcdb.com',
        'port':63440,
        'user':'shendu',
        'passwd':'P@55word',
        'db':'observer_admin',
    }


def process():
    wb = load_workbook(filename='/mnt/f/FileRecv/model.xlsx', read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    l1 = None
    l2 = None
    l3 = None

    for row in ws.rows:
        c3 = row[3].value

        if not c3:
            continue

        c0 = row[0].value
        c1 = row[1].value
        c2 = row[2].value
        c4 = row[4].value
        name = c3
        desc = c4 if c4 else ''

        if c0:
            l1 = c0
            save(sql='INSERT INTO `base_industry` VALUES(%s, %s, %s, %s, %s)', 
                list1=(l1, name, 1, desc, None, ),
                db_config=db_config()
            )

        if c1:
            l2 = c1
            save(sql='INSERT INTO `base_industry` VALUES(%s, %s, %s, %s, %s)', 
                list1=(l2, name, 2, desc, l1, ),
                db_config=db_config()
            )

        if c2:
            l3 = c2
            save(sql='INSERT INTO `base_industry` VALUES(%s, %s, %s, %s, %s)', 
                list1=(l3, name, 3, desc, l2, ),
                db_config=db_config()
            )

        print(name, desc,)


if __name__ == '__main__':
    process()
